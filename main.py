from PyQt6 import QtCore, QtGui, QtWidgets, uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QStackedWidget, QFileDialog
import sys
from database import Database
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import pandas as pd

# Hàm để căn cửa sổ ra giữa màn hình
def center_widget(widget):
    screen = QApplication.primaryScreen()
    screen_geometry = screen.availableGeometry()
    widget_geometry = widget.frameGeometry()
    widget_geometry.moveCenter(screen_geometry.center())
    widget.move(widget_geometry.topLeft())

# Cửa sổ đăng nhập
class DangNhap(QMainWindow):
    def __init__(self):
        super(DangNhap, self).__init__()
        uic.loadUi("DangNhap.ui", self)
        center_widget(self)
        self.btnLogin.clicked.connect(self.DangNhap)

    def DangNhap(self):
        username = self.txtName.text()
        password = self.txtPass.text()
        if username == "1" and password == "1":
            # clear text
            self.txtName.setText("")
            self.txtPass.setText("")
            widget.setFixedWidth(981)
            widget.setFixedHeight(728)
            center_widget(widget)
            widget.setCurrentIndex(1)
        else:
            QMessageBox.information(self, "Thông báo", "Sai tên đăng nhập hoặc mật khẩu")

# Cửa sổ Trang Chủ
class TrangChu(QMainWindow):
    def __init__(self):
        super(TrangChu, self).__init__()
        uic.loadUi("TrangChu.ui", self)
        widget.setFixedWidth(1081)
        widget.setFixedHeight(900)
        center_widget(self)

        self.btnQuanLyNguoiDung.clicked.connect(self.QuanLyNguoiDung)
        self.btnQuanLySanPham.clicked.connect(self.QuanLySanPham)
        self.btnQuanlyDichVu.clicked.connect(self.QuanlyDichVu)
        self.btnQuanLyDonHang.clicked.connect(self.ThongKeDoanhThu)
        self.btnDangXuat.clicked.connect(self.DangXuat)

    def QuanLyNguoiDung(self):
        widget.setFixedWidth(1215)
        widget.setFixedHeight(721)
        center_widget(widget)
        widget.setCurrentIndex(2)

    def QuanLySanPham(self):
        widget.setFixedWidth(1215)
        widget.setFixedHeight(721)
        center_widget(widget)
        widget.setCurrentIndex(3)

    def QuanlyDichVu(self):
        widget.setFixedWidth(1079)
        widget.setFixedHeight(561)
        center_widget(widget)
        widget.setCurrentIndex(4)
    def DangXuat(self):
        widget.setFixedWidth(719)
        widget.setFixedHeight(625)
        center_widget(widget)
        widget.setCurrentIndex(0)
    def ThongKeDoanhThu(self):
        widget.setFixedWidth(1084)
        widget.setFixedHeight(642)
        center_widget(widget)
        widget.setCurrentIndex(7)


# Cửa sổ Quản Lý Người Dùng
class QuanLyNguoiDung(QMainWindow):
    def __init__(self):
        super(QuanLyNguoiDung, self).__init__()
        uic.loadUi("QuanLyNguoiDung.ui", self)
        center_widget(self)
        self.db = Database()
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')

        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget.itemSelectionChanged.connect(self.select_user)

        self.btnThem.clicked.connect(self.Them_user)
        self.btnSua.clicked.connect(self.Sua_user)
        self.btnXoa.clicked.connect(self.Xoa)
        self.btnThoat.clicked.connect(self.QuayLai)
        self.btnTimKiem.clicked.connect(self.timKiem)
        self.btnSuDungDichVu.clicked.connect(self.NguoiDungSuDungDichVu)
        self.btnNguoiDungMuaHang.clicked.connect(self.NguoiDungMuaSanPham)
        self.load_user()


    def load_user(self):
        try:
            data = self.db.read_user()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu phim: {str(e)}")

    def select_user(self):
        selected_row = self.tableWidget.currentRow()
        if selected_row < 0:
            return

        try:
            id = self.tableWidget.item(selected_row, 0).text()
            username = self.tableWidget.item(selected_row, 1).text()
            sdt = self.tableWidget.item(selected_row, 2).text()
            email = self.tableWidget.item(selected_row, 3).text()
            diachi = self.tableWidget.item(selected_row, 4).text()
            luong = self.tableWidget.item(selected_row, 5).text()
            vitri = self.tableWidget.item(selected_row, 6).text()
            quyen = self.tableWidget.item(selected_row, 7).text()
            ngaysinh = self.tableWidget.item(selected_row, 8).text()

            self.txtHoTen.setText(username)
            self.txtDienThoai.setText(sdt)
            self.txtEmail.setText(email)
            self.txtDiaChi.setText(diachi)
            self.txtLuong.setText(luong)
            self.txtViTri.setText(vitri)
            self.txtQuyen.setText(quyen)
            self.dateNamSinh.setDate(QtCore.QDate.fromString(ngaysinh, "yyyy-MM-dd"))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi chọn người dùng: {str(e)}")

    def Them_user(self):
        username = self.txtHoTen.text()
        sdt = self.txtDienThoai.text()
        email = self.txtEmail.text()
        diachi = self.txtDiaChi.text()
        luong = self.txtLuong.text()
        vitri = self.txtViTri.text()
        quyen = self.txtQuyen.text()
        ngaysinh = self.dateNamSinh.date().toString("yyyy-MM-dd")

        # Kiểm tra các trường có rỗng không
        if username == "" or sdt == "" or email == "" or diachi == "" or luong == "" or vitri == "" or quyen == "":
            QMessageBox.information(self, "Thông báo", "Vui lòng nhập đầy đủ thông tin.")
            return

        try:
            self.db.insert_user(username, sdt, email, diachi, luong, vitri, quyen, ngaysinh)
            QMessageBox.information(self, "Thông báo", "Đã thêm người dùng thành công.")

            self.load_user()
        except Exception as e:
            print(e)
            QMessageBox.information(self, "Thông báo", "Đã xảy ra lỗi khi thêm người dùng.")

    def Sua_user(self):
        selected_row = self.tableWidget.currentRow()
        id = self.tableWidget.item(selected_row, 0).text()
        username = self.txtHoTen.text()
        sdt = self.txtDienThoai.text()
        email = self.txtEmail.text()
        diachi = self.txtDiaChi.text()
        luong = self.txtLuong.text()
        vitri = self.txtViTri.text()
        quyen = self.txtQuyen.text()
        ngaysinh = self.dateNamSinh.date().toString("yyyy-MM-dd")
        self.db.update_user(id, username, sdt, email, diachi, luong, vitri, quyen, ngaysinh)
        QMessageBox.information(self, "Thông báo", "Đã cập nhật thông tin thành công.")
        self.load_user()

    def Xoa(self):
        try:
            selected_row = self.tableWidget.currentRow()
            id = self.tableWidget.item(selected_row, 0).text()
            self.db.delete_user(id)
            QMessageBox.information(self, "Thông báo", "Đã xóa người dùng thành công.")
            self.load_user()
        except Exception as e:
            QMessageBox.information(self, "Thông báo", "Đã xảy ra lỗi khi xóa người dùng.")

    def timKiem(self):
        timkiem = self.txtTimKiem.text()
        if timkiem == "":
            self.load_user()
        else:
            try:
                data = self.db.search_user(timkiem)
                self.tableWidget.setRowCount(0)
                for row_number, row_data in enumerate(data):
                    self.tableWidget.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            except Exception as e:
                QMessageBox.warning(self, "Lỗi", f"Lỗi khi tìm kiếm người dùng: {str(e)}")

    def NguoiDungSuDungDichVu(self):
        widget.setFixedWidth(997)
        widget.setFixedHeight(598)
        center_widget(self)
        widget.setCurrentIndex(5)
    def NguoiDungMuaSanPham(self):
        widget.setFixedWidth(1074)
        widget.setFixedHeight(541)
        center_widget(self)
        widget.setCurrentIndex(6)
    def QuayLai(self):
        widget.setCurrentIndex(1)

class QuanLySanPham(QMainWindow):
    def __init__(self):
        super(QuanLySanPham, self).__init__()
        uic.loadUi("QuanLySanPham.ui", self)
        center_widget(self)
        self.db = Database()
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')  # Tìm tableWidget trong UI

        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget.itemSelectionChanged.connect(self.select_row)

        self.btnThem.clicked.connect(self.Them)
        self.btnSua.clicked.connect(self.Sua)
        self.btnXoa.clicked.connect(self.Xoa)
        self.btnThoat.clicked.connect(self.QuayLai)
        self.btnTimKiem.clicked.connect(self.timKiem)
        self.load_product()

    def load_product(self):
        try:
            data = self.db.read_product()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu sản phẩm: {str(e)}")

    def select_row(self):
        selected_row = self.tableWidget.currentRow()
        if selected_row < 0:
            return

        try:
            id = self.tableWidget.item(selected_row, 0).text()
            ten_san_pham = self.tableWidget.item(selected_row, 1).text()
            loai_san_pham = self.tableWidget.item(selected_row, 2).text()
            gia = self.tableWidget.item(selected_row, 3).text()
            mo_ta = self.tableWidget.item(selected_row, 4).text()
            so_luong = self.tableWidget.item(selected_row, 5).text()
            chat_lieu = self.tableWidget.item(selected_row, 6).text()
            da_ban = self.tableWidget.item(selected_row, 7).text()

            self.txtTenSanPham.setText(ten_san_pham)
            self.txtLoaiSanPham.setText(loai_san_pham)
            self.txtGia.setText(gia)
            self.txtMoTa.setText(mo_ta)
            self.txtSoLuongSanPham.setText(so_luong)
            self.txtChatLieu.setText(chat_lieu)
            self.txtDaBan.setText(da_ban)

        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi chọn sản phẩm: {str(e)}")
    def Them(self):
        ten_san_pham = self.txtTenSanPham.text()
        loai_san_pham = self.txtLoaiSanPham.text()
        gia = self.txtGia.text()
        mo_ta = self.txtMoTa.text()
        so_luong = self.txtSoLuongSanPham.text()
        chat_lieu = self.txtChatLieu.text()
        da_ban = self.txtDaBan.text()

        # Kiểm tra các trường có rỗng không
        if ten_san_pham == "" or loai_san_pham == "" or gia == "" or mo_ta == "" or so_luong == "" or chat_lieu == "" or da_ban == "":
            QMessageBox.information(self, "Thông báo", "Vui lòng nhập đầy đủ thông tin.")
            return

        try:
            self.db.insert_product(ten_san_pham, loai_san_pham, gia, mo_ta, so_luong, chat_lieu, da_ban)
            QMessageBox.information(self, "Thông báo", "Đã thêm sản phẩm thành công.")
            self.load_product()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi thêm sản phẩm: {str(e)}")

    def Sua(self):
        selected_row = self.tableWidget.currentRow()
        id = self.tableWidget.item(selected_row, 0).text()
        ten_san_pham = self.txtTenSanPham.text()
        loai_san_pham = self.txtLoaiSanPham.text()
        gia = self.txtGia.text()
        mo_ta = self.txtMoTa.text()
        so_luong = self.txtSoLuongSanPham.text()
        chat_lieu = self.txtChatLieu.text()
        da_ban = self.txtDaBan.text()
        self.db.update_product(id, ten_san_pham, loai_san_pham, gia, mo_ta, so_luong, chat_lieu, da_ban)
        QMessageBox.information(self, "Thông báo", "Đã cập nhật thông tin thành công.")
        self.load_product()

    def Xoa(self):
        try:
            selected_row = self.tableWidget.currentRow()
            id = self.tableWidget.item(selected_row, 0).text()
            self.db.delete_product(id)
            QMessageBox.information(self, "Thông báo", "Đã xóa sản phẩm thành công.")
            self.load_product()
        except Exception as e:
            QMessageBox.information(self, "Thông báo", "Đã xảy ra lỗi khi xóa sản phẩm.")
    def timKiem(self):
        timkiem = self.txtTimKiem.text()
        if timkiem == "":
            self.load_product()
        else:
            try:
                data = self.db.search_product(timkiem)
                self.tableWidget.setRowCount(0)
                for row_number, row_data in enumerate(data):
                    self.tableWidget.insertRow(row_number)
                    for column_number, data in enumerate(row_data):
                        self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            except Exception as e:
                QMessageBox.warning(self, "Lỗi", f"Lỗi khi tìm kiếm sản phẩm: {str(e)}")
    def QuayLai(self):
        widget.setCurrentIndex(1)

class QuanLyDichVu(QMainWindow):
    def __init__(self):
        super(QuanLyDichVu, self).__init__()
        uic.loadUi("QuanLyDichVu.ui", self)
        center_widget(self)
        self.db = Database()
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')  # Tìm tableWidget trong UI

        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget.itemSelectionChanged.connect(self.select_row)

        self.btnThem.clicked.connect(self.Them)
        self.btnSua.clicked.connect(self.Sua)
        self.btnXoa.clicked.connect(self.Xoa)
        self.btnThoat.clicked.connect(self.QuayLai)
        self.load_service()

    def load_service(self):
        try:
            data = self.db.read_service()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu dịch vụ: {str(e)}")

    def select_row(self):
        selected_row = self.tableWidget.currentRow()
        if selected_row < 0:
            return

        try:
            id = self.tableWidget.item(selected_row, 0).text()
            ten_dich_vu = self.tableWidget.item(selected_row, 1).text()
            gia = self.tableWidget.item(selected_row, 2).text()
            mo_ta = self.tableWidget.item(selected_row, 3).text()

            self.txtTenDichVu.setText(ten_dich_vu)
            self.txtGia.setText(gia)
            self.txtMoTa.setText(mo_ta)

        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi chọn dịch vụ: {str(e)}")

    def Them(self):
        ten_dich_vu = self.txtTenDichVu.text()
        gia = self.txtGia.text()
        mo_ta = self.txtMoTa.text()

        # Kiểm tra các trường có rỗng không
        if ten_dich_vu == "" or gia == "" or mo_ta == "":
            QMessageBox.information(self, "Thông báo", "Vui lòng nhập đầy đủ thông tin.")
            return
        try:
            self.db.insert_service(ten_dich_vu, gia, mo_ta)
            QMessageBox.information(self, "Thông báo", "Đã thêm dịch vụ thành công.")
            self.load_service()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi thêm dịch vụ: {str(e)}")



    def Sua(self):
        selected_row = self.tableWidget.currentRow()
        id = self.tableWidget.item(selected_row, 0).text()
        ten_dich_vu = self.txtTenDichVu.text()
        gia = self.txtGia.text()
        mo_ta = self.txtMoTa.text()
        self.db.update_service(id, ten_dich_vu, gia, mo_ta)
        QMessageBox.information(self, "Thông báo", "Đã cập nhật thông tin thành công.")
        self.load_service()

    def Xoa(self):
        try:
            selected_row = self.tableWidget.currentRow()
            id = self.tableWidget.item(selected_row, 0).text()
            self.db.delete_service(id)
            QMessageBox.information(self, "Thông báo", "Đã xóa dịch vụ thành công.")
            self.load_service()
        except Exception as e:
            QMessageBox.information(self, "Thông báo", "Đã xảy ra lỗi khi xóa dịch vụ.")

    def QuayLai(self):
        widget.setFixedWidth(1081)
        widget.setFixedHeight(900)
        center_widget(self)
        widget.setCurrentIndex(1)

class NguoiDungMuaSanPham(QMainWindow):
    def __init__(self):
        super(NguoiDungMuaSanPham, self).__init__()

        uic.loadUi("NguoiDungMuaSanPham.ui", self)
        center_widget(self)
        self.db = Database()
        self.id_don_hang = None
        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')
        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget.itemSelectionChanged.connect(self.select_row)
        self.btnThemDonHang.clicked.connect(self.ThemDonHang)
        self.btnThemSanPham.clicked.connect(self.ThemSanPham)
        self.cboSanPham.currentIndexChanged.connect(self.on_combobox_changed)
        self.btnSuaDonHang.clicked.connect(self.SuaTrangThaiDonHang)
        self.btnXoaDonHang.clicked.connect(self.XoaDonHang)
        self.btnSuaSanPham.clicked.connect(self.CapNhatSoLuongSanPham)
        self.btnThoatDonHang.clicked.connect(self.QuayLai)
        self.loading()
    def loading(self):
        try:
            select_user = self.db.select_user()
            select_product = self.db.read_product()
            all_order = self.db.select_tt_order_product()
            # clear combobox
            self.cboNguoiDung.clear()
            self.cboSanPham.clear()
            for product in select_product:
                self.cboSanPham.addItem(product[1])
            for user in select_user:
                self.cboNguoiDung.addItem(user[1])
            # loading all_order len table widget
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(all_order):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))

        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu người dùng: {str(e)}")
    def ThemDonHang(self):
        try:
            # lay ten user tu comboBox roi lay id tuong ung
            id_user = self.db.select_user()[self.cboNguoiDung.currentIndex()][0]
            trang_thai_don_hang = self.cboTrangThaiDongHang.currentText()
            self.id_don_hang = self.db.insert_order(id_user, trang_thai_don_hang)[0]
            QMessageBox.information(self, "Thông báo", "Đã tạo đơn hàng thành công.")
            self.loading()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tạo đơn hàng: {str(e)}")
    def ThemSanPham(self):
        try:
            # lay ten san pham tu comboBox roi lay id tuong ung
            id_product = self.db.read_product()[self.cboSanPham.currentIndex()][0]
            print("check id_product", id_product)
            so_luong = self.txtSoLuong.text()
            # kiem tra so luong co rong khong
            if so_luong == "":
                QMessageBox.information(self, "Thông báo", "Vui lòng nhập số lượng.")
                return
            if not so_luong.isdigit():
                QMessageBox.information(self, "Thông báo", "Số lượng phải là số.")
                return
            if int(so_luong) <= 0:
                QMessageBox.information(self, "Thông báo", "Số lượng phải lớn hơn 0.")
                return
            self.db.insert_product_order(id_product, self.id_don_hang, so_luong)
            self.loading()
            QMessageBox.information(self, "Thông báo", "Đã thêm sản phẩm vào đơn hàng thành công.")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi thêm sản phẩm vào đơn hàng: {str(e)}")

    # su kien khi select cboSanPham se hien thi so luong san pham tuong ung neu co

    def select_row(self):
        selected_items = self.tableWidget.selectedItems()
        if selected_items:
            try:
                #lay trang thai don hang
                trang_thai_don_hang = self.tableWidget.item(selected_items[0].row(), 5).text()
                ten_san_pham =  self.tableWidget.item(selected_items[0].row(), 6).text()
                so_luong = self.tableWidget.item(selected_items[0].row(), 11).text()
                # set trang thai don hang tren comboBox tuong ung
                self.cboTrangThaiDongHang.setCurrentText(trang_thai_don_hang)
                # set ten san pham tren comboBox tuong ung
                self.cboSanPham.setCurrentText(ten_san_pham)
                self.txtSoLuong.setText(so_luong)
            except Exception as e:
                QMessageBox.warning(self, "Lỗi", f"Lỗi khi chọn đơn hàng: {str(e)}")

    def on_combobox_changed(self):
        # Kiểm tra combobox nào đã thay đổi
        changed_combobox = self.sender()

        # Nếu combobox thay đổi là cboSanPham thì thực hiện xử lý
        if changed_combobox == self.cboSanPham:
            try:
                id_product = self.db.read_product()[self.cboSanPham.currentIndex()][0]
                quantity = self.db.select_quantity_product_order(self.id_don_hang, id_product)[0]

                if quantity:
                    self.txtSoLuong.setText(str(quantity))
                else:
                    self.txtSoLuong.setText("Chưa mua")
            except Exception as e:
                self.txtSoLuong.setText("0")
    def SuaTrangThaiDonHang(self):
        try:
            trang_thai_don_hang = self.cboTrangThaiDongHang.currentText()
            # lay id don hang tu table widget
            selected_items = self.tableWidget.selectedItems()
            if selected_items:
               id_dh = self.tableWidget.item(selected_items[0].row(), 0).text()
            # check xem co id don hang nao chua
            print("Check id_dh", id_dh)
            print("Check trang thai don hang", trang_thai_don_hang)
            if id_dh == "":
                QMessageBox.information(self, "Thông báo", "Vui lòng chọn đơn hàng.")
            else:
                self.db.cap_nhat_trang_thai_don_hang(id_dh, trang_thai_don_hang)
                QMessageBox.information(self, "Thông báo", "Đã cập nhật trạng thái đơn hàng thành công.")
                self.loading()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi cập nhật trạng thái đơn hàng: {str(e)}")
    def XoaDonHang(self):
        try:
            selected_items = self.tableWidget.selectedItems()
            if selected_items:
                id_dh = self.tableWidget.item(selected_items[0].row(), 0).text()
                if id_dh == "":
                    QMessageBox.information(self, "Thông báo", "Vui lòng chọn đơn hàng.")
                else:
                    self.db.xoa_don_hang(id_dh)
                    QMessageBox.information(self, "Thông báo", "Đã xóa đơn hàng thành công.")
                    self.loading()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi xóa đơn hàng: {str(e)}")
    def CapNhatSoLuongSanPham(self):
        try:
            # lay id san pham tu comboBox
            id_product = self.db.read_product()[self.cboSanPham.currentIndex()][0]
            so_luong = self.txtSoLuong.text()
            # kiem tra so luong co rong khong
            if so_luong == "":
                QMessageBox.information(self, "Thông báo", "Vui lòng nhập số lượng.")
                return
            if not so_luong.isdigit():
                QMessageBox.information(self, "Thông báo", "Số lượng phải là số.")
                return
            if int(so_luong) <= 0:
                QMessageBox.information(self, "Thông báo", "Số lượng phải lớn hơn 0.")
                return
            # lay id don hang tu table widget
            selected_items = self.tableWidget.selectedItems()
            if selected_items:
                id_dh = self.tableWidget.item(selected_items[0].row(), 0).text()
            # check xem co id don hang nao chua
            if id_dh == "":
                QMessageBox.information(self, "Thông báo", "Vui lòng chọn đơn hàng.")
            else:
                self.db.cap_nhat_so_luong_product_order(id_product, id_dh, so_luong)
                QMessageBox.information(self, "Thông báo", "Đã cập nhật số lượng sản phẩm thành công.")
                self.loading()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi cập nhật số lượng sản phẩm: {str(e)}")
    def QuayLai(self):
        widget.setFixedWidth(1081)
        widget.setFixedHeight(900)
        center_widget(self)
        widget.setCurrentIndex(1)

class NguoiDungMuaDichVu(QMainWindow):
    def __init__(self):
        super(NguoiDungMuaDichVu, self).__init__()
        uic.loadUi("NguoiDungSuDungDichVu.ui", self)
        center_widget(self)
        self.db = Database()
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')  # Tìm tableWidget trong UI
        # Kết nối sự kiện khi chọn một dòng trong tableWidget
        self.tableWidget.itemSelectionChanged.connect(self.select_row)
        self.btnAdd.clicked.connect(self.Them)
        self.btnUpdate.clicked.connect(self.Sua)
        self.btnDelete.clicked.connect(self.Xoa)
        self.btnClose.clicked.connect(self.QuayLai)
        self.load_service()

    def load_service(self):
        try:
            select_user = self.db.select_user()
            # clear combobox
            self.cboThongTinNguoiDung.clear()
            for user in select_user:
                self.cboThongTinNguoiDung.addItem(user[1])
            data = self.db.read_service()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu dịch vụ: {str(e)}")

    def select_row(self):
        selected_items = self.tableWidget.selectedItems()
        if selected_items:
            try:
               #lay du lieu dong da select
                id_dich_vu = self.tableWidget.item(selected_items[0].row(), 0).text()
                id_user = self.db.select_user()[self.cboThongTinNguoiDung.currentIndex()][0]

                so_luong = self.db.select_quantity_user_service(id_user, id_dich_vu)[0]
                if so_luong:
                    self.txtSoLuong.setText(str(so_luong))
                else:
                    self.txtSoLuong.setText("Chưa sử dụng")
            except Exception as e:
                so_luong = "Chưa sử dụng"
                self.txtSoLuong.setText(so_luong)
                return
    def Them(self):
        #neu chua chon dich vu nào ma them thi thong bao
        if self.tableWidget.currentRow() == -1:
            QMessageBox.information(self, "Thông báo", "Vui lòng chọn dịch vụ.")
            return
        d_dich_vu = self.tableWidget.item(self.tableWidget.currentRow(), 0).text()
        # neu chon tren comboBox id nguoi dung thi lay id nguoi dung tuong ung luc
        # loadding len
        ten_nguoi_dung = self.cboThongTinNguoiDung.currentIndex()
        id_nguoi_dung = self.db.select_user()[ten_nguoi_dung][0]
        so_luong = self.txtSoLuong.text()
        # kiem tra so luong co rong khong
        if so_luong == "":
            QMessageBox.information(self, "Thông báo", "Vui lòng nhập số lượng.")
            return
        if not so_luong.isdigit():
            QMessageBox.information(self, "Thông báo", "Số lượng phải là số.")
            return
        if int(so_luong) <= 0:
            QMessageBox.information(self, "Thông báo", "Số lượng phải lớn hơn 0.")
            return
        try:
            self.db.insert_user_service(id_nguoi_dung, d_dich_vu, so_luong)
            QMessageBox.information(self, "Thông báo", "Đã thêm dịch vụ thành công.")
            self.txtSoLuong.setText("")
            self.load_service()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi thêm dịch vụ: {str(e)}")


    def Sua(self):
        # validate va co thong bao
        ten_user = self.cboThongTinNguoiDung.currentText()
        id_user = self.db.select_user()[self.cboThongTinNguoiDung.currentIndex()][0]
        id_service = self.tableWidget.item(self.tableWidget.currentRow(), 0).text()
        so_luong = self.txtSoLuong.text()
        # validate
        if so_luong == "":
            QMessageBox.information(self, "Thông báo", "Vui lòng nhập số lượng.")
            return
        if not so_luong.isdigit():
            QMessageBox.information(self, "Thông báo", "Số lượng phải là số.")
            return
        if int(so_luong) <= 0:
            QMessageBox.information(self, "Thông báo", "Số lượng phải lớn hơn 0.")
            return
        try:
            self.db.update_user_service(id_user, id_service, so_luong)
            QMessageBox.information(self, "Thông báo", "Đã cập nhật thông tin thành công.")
            self.load_service()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi cập dữ liệu: {str(e)}")
    def Xoa(self):
        ten_user = self.cboThongTinNguoiDung.currentText()
        id_user = self.db.select_user()[self.cboThongTinNguoiDung.currentIndex()][0]
        id_service = self.tableWidget.item(self.tableWidget.currentRow(), 0).text()
        try:
            self.db.delete_user_service(id_user, id_service)
            QMessageBox.information(self, "Thông báo", "Đã xóa người dùng sử dụng dịch vụ thành công.")
            self.load_service()
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi xóa: {str(e)}")
    def QuayLai(self):
        widget.setFixedWidth(1081)
        widget.setFixedHeight(900)
        center_widget(self)
        widget.setCurrentIndex(1)


class ThongKeDoanhThu(QMainWindow):
    def __init__(self):
        super(ThongKeDoanhThu, self).__init__()
        uic.loadUi("ThongKeDoanhThu.ui", self)
        center_widget(self)
        self.db = Database()
        self.tongSoTien = 0
        self.soLuong = 0
        self.btnTheoDichVu.clicked.connect(self.TheoDichVu)
        self.btnTheoSanPham.clicked.connect(self.TheoSanPham)
        self.btnInBaoCao.clicked.connect(self.InBaoCao)
        self.btnThoat.clicked.connect(self.QuayLai)
        # table widget
        self.tableWidget = self.findChild(QtWidgets.QTableWidget, 'tableWidget')

    def TheoSanPham(self):
        self.tableWidget.setColumnCount(13)
        self.tableWidget.setHorizontalHeaderLabels(
            ["ID", "Tên", "Số điện thoại", "Email", "Địa chỉ", "Trạng thái đơn hàng", "Tên", "Loại", "Mô tả",
             "Chất liệu", "Giá", "Số lượng", "Số tiền phải trả"])
        try:
            self.tongSoTien = 0
            self.soLuong = 0
            data = self.db.select_tt_order_product()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.soLuong += 1
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    # tinh tong so tien neu trạng thái đơn hàng là Hoàn tất
                    if column_number == 12 and row_data[5] == "Hoàn tất":
                        self.tongSoTien += data
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            # hien thi tong so tien
            # fort mat thanh tien Viet Nam  1.000.000
            self.txtTongTien.setText("{:,}".format(self.tongSoTien) + "VNĐ")
            self.txtSoLuongDon.setText(str(self.soLuong))
            # self.txtTongTien.setText(str(self.tongSoTien))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu sản phẩm: {str(e)}")
    def TheoDichVu(self):
        self.tongSoTien = 0
        self.soLuong = 0
        self.tableWidget.setColumnCount(11)
        self.tableWidget.setHorizontalHeaderLabels(
            ["ID", "Tên", "Số điện thoại", "Email", "Địa chỉ", "Ngày sinh", "Tên dịch vụ", "Giá", "Mô tả",
             "Số lượng", "Số tiền phải trả"])
        try:
            data = self.db.select_tt_order_service()
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(data):
                self.soLuong += 1
                self.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    # tinh tong so tien
                    if column_number == 10:
                        self.tongSoTien += data
                    self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            # hien thi tong so tien
            self.txtTongTien.setText("{:,}".format(self.tongSoTien) + "VNĐ")
            self.txtSoLuongDon.setText(str(self.soLuong))
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi tải dữ liệu dịch vụ: {str(e)}")

    def InBaoCao(self):
        try:
            file_name = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
            if file_name[0]:
                wb = Workbook()
                ws = wb.active

                # Định dạng font in đậm và đường viền
                bold_font = Font(bold=True)
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                # Thêm tiêu đề vào trước
                if self.tableWidget.columnCount() == 13:
                    headers = ["ID", "Tên", "Số điện thoại", "Email", "Địa chỉ", "Trạng thái đơn hàng", "Tên", "Loại",
                               "Mô tả", "Chất liệu", "Giá", "Số lượng", "Số tiền phải trả"]
                elif self.tableWidget.columnCount() == 11:
                    headers = ["ID", "Tên", "Số điện thoại", "Email", "Địa chỉ", "Ngày sinh", "Tên dịch vụ", "Giá",
                               "Mô tả", "Số lượng", "Số tiền phải trả"]

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = bold_font
                    cell.border = thin_border

                # Sau đó thêm dữ liệu
                for row in range(self.tableWidget.rowCount()):
                    for col in range(self.tableWidget.columnCount()):
                        cell = ws.cell(row=row + 2, column=col + 1, value=self.tableWidget.item(row, col).text())
                        cell.border = thin_border  # Thêm đường viền cho các ô dữ liệu

                # Thêm dòng tổng kết
                ws.append(["Tổng số lượng đơn hàng: ", str(self.soLuong)])
                ws.append(["Tổng số tiền", "{:,}".format(self.tongSoTien) + "VNĐ"])

                # Định dạng đường viền cho dòng tổng kết
                last_row = ws.max_row
                for col_num in range(1, self.tableWidget.columnCount() + 1):
                    cell = ws.cell(row=last_row - 1, column=col_num)
                    cell.border = thin_border
                    cell = ws.cell(row=last_row, column=col_num)
                    cell.border = thin_border

                wb.save(file_name[0])
                QMessageBox.information(self, "Thông báo", "Đã lưu báo cáo thành công.")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi khi lưu báo cáo: {str(e)}")
    def QuayLai(self):
        widget.setFixedWidth(1081)
        widget.setFixedHeight(900)
        center_widget(self)
        widget.setCurrentIndex(1)

# Khởi tạo ứng dụng
app = QApplication(sys.argv)
widget = QtWidgets.QStackedWidget()

# Khởi tạo các cửa sổ
dangnhap_f = DangNhap()
trangchu_f = TrangChu()
quanlynguoidung_f = QuanLyNguoiDung()
quanlysanpham_f = QuanLySanPham()
quanlydichvu_f = QuanLyDichVu()
nguoidungmuasanpham_f = NguoiDungMuaSanPham()
nguoidungmuadichvu_f = NguoiDungMuaDichVu()
thongkedoanhthu_f = ThongKeDoanhThu()
# Thêm các cửa sổ vào stacked widget
widget.addWidget(dangnhap_f)
widget.addWidget(trangchu_f)
widget.addWidget(quanlynguoidung_f)
widget.addWidget(quanlysanpham_f)
widget.addWidget(quanlydichvu_f)
widget.addWidget(nguoidungmuasanpham_f)
widget.addWidget(nguoidungmuadichvu_f)
widget.addWidget(thongkedoanhthu_f)


# Thiết lập cửa sổ ban đầu
widget.setCurrentIndex(0)
widget.setFixedWidth(719)
widget.setFixedHeight(625)
center_widget(widget)
widget.show()

sys.exit(app.exec())


